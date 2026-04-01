#!/usr/bin/env python3
import argparse, json, datetime
from openpyxl import Workbook, load_workbook

COLUMNS = [
    "Timestamp",
    "Source",
    "Company",
    "Role title",
    "Location",
    "Link to posting",
    "ATS fit score",
    "Keywords matched",
    "Applied status",
    "Resume path",
    "Cover letter path",
    "Application portal",
    "Notes",
    "Follow up date suggestion",
    "Blocker reason",
]

def ensure_tracker(path: str):
    try:
        wb = load_workbook(path)
        ws = wb.active
        if ws.max_row == 0:
            raise Exception("empty")
        return wb, ws
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.append(COLUMNS)
        wb.save(path)
        return wb, ws

def add_row(tracker: str, job_json_path: str):
    wb, ws = ensure_tracker(tracker)
    job = json.load(open(job_json_path, "r", encoding="utf-8"))
    row = [
        datetime.datetime.now().isoformat(timespec="seconds"),
        job.get("source",""),
        job.get("company",""),
        job.get("role_title",""),
        job.get("location",""),
        job.get("posting_url",""),
        job.get("ats_score",""),
        ", ".join(job.get("keywords_matched", [])) if isinstance(job.get("keywords_matched"), list) else job.get("keywords_matched",""),
        job.get("status",""),
        job.get("resume_path",""),
        job.get("cover_letter_path",""),
        job.get("portal",""),
        job.get("notes",""),
        job.get("follow_up_suggestion",""),
        job.get("blocker_reason",""),
    ]
    ws.append(row)
    wb.save(tracker)

def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    add = sub.add_parser("add")
    add.add_argument("--tracker", required=True)
    add.add_argument("--job_json", required=True)
    args = ap.parse_args()
    if args.cmd == "add":
        add_row(args.tracker, args.job_json)

if __name__ == "__main__":
    main()
