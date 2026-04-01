#!/usr/bin/env python3
import argparse
import datetime
import json
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook


COLUMNS = [
    "Timestamp",
    "Source",
    "Company",
    "Role title",
    "Location",
    "Link to posting",
    "ATS fit score",
    "Keywords matched",
    "Applied status",
    "Resume path",
    "Cover letter path",
    "Application portal",
    "Notes",
    "Follow up date suggestion",
    "Blocker reason",
]


def default_run_root() -> Path:
    env_root = os.environ.get("JOB_APPLY_RUN_ROOT")
    if env_root:
        return Path(env_root).expanduser()
    return Path.home() / "Desktop" / "Job Apply Runs" / "current"


def resolve_artifacts_root(run_root: Path) -> Path:
    if (run_root / "jobs").exists() and (run_root / "job_tracker.xlsx").exists():
        return run_root
    return run_root / "artifacts"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Initialize or append an XLSX job tracker from job folders.")
    parser.add_argument("--run-root", default=str(default_run_root()), help="Run root or artifacts root.")
    parser.add_argument("--jobs-dir", default="", help="Explicit jobs directory override.")
    parser.add_argument("--tracker-path", default="", help="Explicit tracker path override.")
    return parser.parse_args()


def ensure_tracker(path: Path):
    if path.exists():
        return load_workbook(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.append(COLUMNS)
    wb.save(path)
    return wb


def main() -> int:
    args = parse_args()
    run_root = Path(args.run_root).expanduser()
    artifacts_root = resolve_artifacts_root(run_root)

    jobs_dir = Path(args.jobs_dir).expanduser() if args.jobs_dir else artifacts_root / "jobs"
    tracker_path = Path(args.tracker_path).expanduser() if args.tracker_path else artifacts_root / "job_tracker.xlsx"

    if not jobs_dir.exists():
        print(f"Jobs directory not found: {jobs_dir}")
        return 2

    wb = ensure_tracker(tracker_path)
    ws = wb.active

    job_dirs = [d for d in sorted(jobs_dir.iterdir()) if d.is_dir()]
    print(f"Adding {len(job_dirs)} jobs to tracker...")

    added = 0
    for job_dir in job_dirs:
        job_json_path = job_dir / "job.json"
        resume_path = job_dir / "resume.pdf"
        if not job_json_path.exists():
            continue

        job = json.loads(job_json_path.read_text(encoding="utf-8"))
        role = job.get("role", job.get("role_title", "Unknown"))
        company = job.get("company", "Unknown")
        location = job.get("location", "")
        url = job.get("url", job.get("posting_url", ""))

        ws.append(
            [
                datetime.datetime.now().isoformat(timespec="seconds"),
                "LinkedIn",
                company,
                role,
                location,
                url,
                "",
                "",
                "Captured",
                str(resume_path),
                "",
                "",
                "",
                "",
                "",
            ]
        )
        added += 1

    wb.save(tracker_path)
    print(f"Tracker updated at {tracker_path} ({added} rows added)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
